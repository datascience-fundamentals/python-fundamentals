import openpyxl

wb = openpyxl.load_workbook("doc/planilla.xlsx")
print(wb.sheetnames)  # showing the excel sheet names
# print(wb["Hoja1"]) # selecting an excel sheet for its name

sheet = wb.active  # selecting the active excel sheet
print(sheet)

wb.create_sheet("Hoja3")  # creating an excel sheet
print(wb.sheetnames)

sheet3 = wb["Hoja3"]
sheet3.title = "Nuevo titulo"  # changing the excel sheet name
print(wb.sheetnames)

print(
    sheet.max_row,  # 4
    sheet.max_column,  # 3
)

cell = sheet["A1"]  # selecting a cell
print(cell.value)  # showing the cell value
cell.value = "Nombre completo"
print(cell.value)  # changing the cell value


# selecting a cell by specifying it sposition
cell2 = sheet.cell(row=2, column=1)
print(
    cell2.value,  # the cell value
    cell2.row,  # the cell row position
    cell2.column,  # the cell column position
    cell2.coordinate,  # the cell coordinate value since the excel document
)

# iterating cells from the excel document
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(row, column, cell.value)

# selecting an excel column
column = sheet["A"]
# getting a cell value from the excel column
print(column[0].value)

# selecting an excel row
row = sheet["1"]
# getting a cell value from the excel row
print(row[0].value)

# adding a new excel row with 3 columns
sheet.append([1, 2, 3])
print(sheet.max_row)

sheet.delete_rows(
    1,  # deleting the first row
    1,  # number of rows to delete
)
print(sheet.max_row)

wb.save("doc/new_excel.xlsx")
